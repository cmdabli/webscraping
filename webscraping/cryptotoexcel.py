import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cryptocurrencies"

header_font = Font(bold=True, color="FFFFFF")
ws["A1"].font = header_font
ws["B1"].font = header_font
ws["C1"].font = header_font
ws["D1"].font = header_font

ws["A1"] = "Name"
ws["B1"] = "Symbol"
ws["C1"] = "Current Price"
ws["D1"] = "% Change (24hr)"

ws["A2"] = "Bitcoin"
ws["B2"] = "BTC"
ws["C2"] = "$29,361.78"
ws["D2"] = "0.90%"

ws["A3"] = "Ethereum"
ws["B3"] = "ETH"
ws["C3"] = "$1,950.95"
ws["D3"] = "1.50%"

ws["A4"] = "Tether"
ws["B4"] = "USDT"
ws["C4"] = "$1.00"
ws["D4"] = "-0.20%"

ws["A5"] = "BNB"
ws["B5"] = "BNB"
ws["C5"] = "$327.28"
ws["D5"] = "0.70%"

ws["A6"] = "USD Coin"
ws["B6"] = "USDC"
ws["C6"] = "$1.00"
ws["D6"] = "0.00%"

for row in ws.iter_rows(min_row=2, max_row=6, min_col=3, max_col=3):
    for cell in row:
        price = cell.value.replace("$", "").replace(",", "")
        price = float(price)
        change = cell.offset(column=1).value.replace("%", "")
        change = float(change) / 100
        corresponding_price = price / (1 + change)
        cell.offset(column=1).value = "{:.2%}".format(change)
        cell.offset(column=2).value = "${:,.2f}".format(corresponding_price)

for row in ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
        cell.font = Font(bold=True, size=14, color="FFFFFF")

for row in ws.iter_rows(min_row=2, max_row=6):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="E9F1F7", end_color="E9F1F7", fill_type="solid")
        cell.font = Font(size=12)

ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15

wb.save("cryptocurrencies.xlsx")



